"""
Bot Keuangan & Produksi Telegram
"""

from keep_alive import keep_alive
keep_alive()

import os
import re
import json
import logging
import datetime as dt
from zoneinfo import ZoneInfo
from io import BytesIO

import gspread
from google.oauth2.service_account import Credentials
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
#  KONFIGURASI
# ──────────────────────────────────────────────
TOKEN       = os.environ["TELEGRAM_TOKEN"]
SHEET_ID    = os.environ["GOOGLE_SHEET_ID"]
CREDS_JSON  = os.environ["GOOGLE_CREDS_JSON"]
OWNER_ID    = int(os.environ["OWNER_ID"])
TIMEZONE    = os.environ.get("TIMEZONE", "Asia/Jakarta")
JAM_LAPORAN = int(os.environ.get("JAM_LAPORAN", "21"))

TZ = ZoneInfo(TIMEZONE)

logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────
#  GOOGLE SHEETS
# ──────────────────────────────────────────────
def get_sheet():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    info  = json.loads(CREDS_JSON)
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    gc    = gspread.authorize(creds)
    return gc.open_by_key(SHEET_ID)


def ensure_sheets(sh):
    existing = [ws.title for ws in sh.worksheets()]

    def make(name, headers, hex_color="6B3F1F"):
        if name not in existing:
            ws = sh.add_worksheet(name, rows=5000, cols=len(headers))
            ws.append_row(headers)
            r = int(hex_color[0:2], 16) / 255
            g = int(hex_color[2:4], 16) / 255
            b = int(hex_color[4:6], 16) / 255
            col_end = get_column_letter(len(headers))
            ws.format(f"A1:{col_end}1", {
                "textFormat": {
                    "bold": True,
                    "foregroundColor": {"red": 1, "green": 1, "blue": 1}
                },
                "backgroundColor": {"red": r, "green": g, "blue": b}
            })

    make("Transaksi", ["ID","Tanggal","Waktu","Tipe","Kategori","Deskripsi","Nominal","Sumber"], "1E7A4A")
    make("Produksi",  ["ID","Tanggal","Waktu","Produk","Jumlah","Satuan","Catatan"], "B7770D")


def get_ws(sh, name):
    return sh.worksheet(name)


# ──────────────────────────────────────────────
#  UTILITAS
# ──────────────────────────────────────────────
def fmt(n):
    return "Rp {:,}".format(int(n)).replace(",", ".")


def now_tz():
    return dt.datetime.now(TZ)


def today_str():
    return now_tz().strftime("%Y-%m-%d")


def this_month():
    return now_tz().strftime("%Y-%m")


def get_next_id(ws):
    vals = ws.col_values(1)
    nums = [int(v) for v in vals[1:] if str(v).isdigit()]
    return (max(nums) + 1) if nums else 1


def esc(text):
    """Escape karakter khusus Markdown agar tidak rusak format pesan."""
    chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    for c in chars:
        text = text.replace(c, f'\\{c}')
    return text


def parse_nominal(s):
    s = s.lower().strip().replace(",", "")
    if "jt" in s or "juta" in s:
        num = re.sub(r"[^\d.]", "", s.replace("jt", "").replace("juta", ""))
        return int(float(num) * 1_000_000)
    if "rb" in s or "ribu" in s or "k" in s:
        num = re.sub(r"[^\d.]", "", s.replace("rb","").replace("ribu","").replace("k",""))
        return int(float(num) * 1_000)
    return int(re.sub(r"[^\d]", "", s))


def auto_kategori(teks, tipe):
    t = teks.lower()
    if tipe == "KELUAR":
        peta = {
            "makan":      ["makan","minum","kopi","snack","sarapan","warung","resto","indomaret","alfamart","jajan","mie","nasi","bakso","soto"],
            "transport":  ["bensin","grab","gojek","ojek","parkir","tol","bus","motor","bensin"],
            "bahan_baku": ["tepung","gula","mentega","telur","susu","coklat","keju","bahan","beli"],
            "operasional":["listrik","air","gas","internet","wifi","pulsa","sewa"],
            "gaji":       ["gaji","upah","honor","tunjangan"],
            "peralatan":  ["alat","peralatan","mesin","cetakan","loyang","oven"],
        }
        for kat, kws in peta.items():
            if any(kw in t for kw in kws):
                return kat
        return "lain_keluar"
    else:
        peta = {
            "penjualan": ["jual","penjualan","laku","omset","hasil"],
            "transfer":  ["transfer","kirim","tf"],
            "modal":     ["modal","investasi","setor"],
        }
        for kat, kws in peta.items():
            if any(kw in t for kw in kws):
                return kat
        return "lain_masuk"


# ──────────────────────────────────────────────
#  GUARD OWNER
# ──────────────────────────────────────────────
def owner_only(func):
    async def wrapper(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != OWNER_ID:
            await update.message.reply_text("Bot ini bersifat privat.")
            return
        await func(update, ctx)
    wrapper.__name__ = func.__name__
    return wrapper


# ──────────────────────────────────────────────
#  PARSER TRANSAKSI
# ──────────────────────────────────────────────
def parse_transaksi(text):
    text = text.strip()
    tipe = "MASUK" if text.startswith("+") else "KELUAR"
    if tipe == "MASUK":
        text = text[1:].strip()

    match = re.search(r'(\d[\d.]*(?:rb|ribu|k|jt|juta)?)\s*$', text, re.IGNORECASE)
    if not match:
        return None

    nominal_str = match.group(1)
    deskripsi   = text[:match.start()].strip()
    if not deskripsi:
        return None

    try:
        nominal = parse_nominal(nominal_str)
    except Exception:
        return None

    if nominal <= 0:
        return None

    return {"tipe": tipe, "deskripsi": deskripsi, "nominal": nominal}


# ──────────────────────────────────────────────
#  PARSER PRODUKSI
# ──────────────────────────────────────────────
def parse_produksi(text):
    text = re.sub(r'^prod(uksi)?\s*', '', text.strip(), flags=re.IGNORECASE).strip()
    items = []
    for part in re.split(r'[,;]', text):
        part = part.strip()
        if not part:
            continue
        m = re.search(r'(\d+)\s*([a-zA-Z]*)\s*$', part)
        if not m:
            continue
        jumlah = int(m.group(1))
        satuan = m.group(2).strip() if m.group(2).strip() else "pcs"
        nama   = part[:m.start()].strip()
        if nama and jumlah > 0:
            items.append({"produk": nama, "jumlah": jumlah, "satuan": satuan})
    return items if items else None


# ──────────────────────────────────────────────
#  RINGKASAN HELPER
# ──────────────────────────────────────────────
def hitung(records, fn=None):
    data   = [r for r in records if fn(r)] if fn else records
    masuk  = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "MASUK")
    keluar = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "KELUAR")
    return masuk, keluar, masuk - keluar


# ──────────────────────────────────────────────
#  COMMAND: /start
# ──────────────────────────────────────────────
@owner_only
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sh = get_sheet()
    ensure_sheets(sh)
    nama = update.effective_user.first_name
    await update.message.reply_text(
        f"Halo {nama}! Bot Keuangan & Produksi aktif\n\n"
        "Ketik /menu untuk melihat semua perintah."
    )


# ──────────────────────────────────────────────
#  COMMAND: /menu
# ──────────────────────────────────────────────
@owner_only
async def cmd_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    teks = (
        "DAFTAR PERINTAH LENGKAP\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "CATAT PENGELUARAN\n"
        "Langsung ketik tanpa perintah:\n"
        "beli tepung 50000\n"
        "bayar listrik 150rb\n"
        "jajan indomaret 93.700\n\n"
        "CATAT PEMASUKAN\n"
        "Awali dengan tanda +\n"
        "+ penjualan roti 250000\n"
        "+ transfer Bu Sari 500rb\n\n"
        "CATAT PRODUKSI\n"
        "Awali dengan produksi atau prod:\n"
        "produksi roti coklat 50\n"
        "prod croissant 30 loyang, bolu 20\n\n"
        "LAPORAN\n"
        "/hari - Laporan hari ini\n"
        "/kemarin - Laporan kemarin\n"
        "/bulan - Laporan bulan ini\n"
        "/saldo - Ringkasan saldo lengkap\n"
        "/labarugi - Laporan laba rugi\n"
        "/prodhari - Produksi hari ini\n"
        "/prodbul - Produksi bulan ini\n\n"
        "EXPORT EXCEL\n"
        "/excel - Excel bulan ini\n"
        "/excelbulan 2025-01 - Bulan tertentu\n\n"
        "HAPUS\n"
        "/hapus - Hapus transaksi terakhir\n"
        "/hapusprod - Hapus produksi terakhir\n\n"
        "FORMAT NOMINAL\n"
        "50000 = Rp 50.000\n"
        "93.700 = Rp 93.700\n"
        "50rb atau 50k = Rp 50.000\n"
        "1.5jt = Rp 1.500.000\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Laporan otomatis dikirim tiap malam"
    )
    await update.message.reply_text(teks)


# ──────────────────────────────────────────────
#  COMMAND: /saldo
# ──────────────────────────────────────────────
@owner_only
async def cmd_saldo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sh      = get_sheet()
    ws      = get_ws(sh, "Transaksi")
    records = ws.get_all_records()
    bulan   = this_month()
    tgl     = today_str()

    m_all, k_all, s_all    = hitung(records)
    m_bul, k_bul, l_bul    = hitung(records, lambda r: str(r["Tanggal"]).startswith(bulan))
    m_hari, k_hari, l_hari = hitung(records, lambda r: r["Tanggal"] == tgl)

    await update.message.reply_text(
        f"RINGKASAN KEUANGAN\n\n"
        f"Hari ini ({tgl})\n"
        f"  Masuk : {fmt(m_hari)}\n"
        f"  Keluar: {fmt(k_hari)}\n"
        f"  Laba  : {fmt(l_hari)}\n\n"
        f"Bulan ini ({bulan})\n"
        f"  Masuk : {fmt(m_bul)}\n"
        f"  Keluar: {fmt(k_bul)}\n"
        f"  Laba  : {fmt(l_bul)}\n\n"
        f"Saldo Total\n"
        f"  {fmt(s_all)}"
    )


# ──────────────────────────────────────────────
#  LAPORAN PER TANGGAL
# ──────────────────────────────────────────────
async def laporan_tanggal(update, tgl_str, label):
    sh      = get_sheet()
    ws      = get_ws(sh, "Transaksi")
    records = ws.get_all_records()
    data    = [r for r in records if r["Tanggal"] == tgl_str]

    masuk  = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "MASUK")
    keluar = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "KELUAR")

    kat_totals = {}
    for r in data:
        if r["Tipe"] == "KELUAR":
            kat = r.get("Kategori", "lain")
            kat_totals[kat] = kat_totals.get(kat, 0) + int(r["Nominal"])

    teks  = f"Laporan {label} ({tgl_str})\n\n"
    teks += f"Total Masuk : {fmt(masuk)}\n"
    teks += f"Total Keluar: {fmt(keluar)}\n"
    teks += f"Laba Bersih : {fmt(masuk - keluar)}\n"
    teks += f"Jumlah TX   : {len(data)}\n"

    if kat_totals:
        teks += "\nPengeluaran per Kategori:\n"
        for kat, total in sorted(kat_totals.items(), key=lambda x: -x[1]):
            teks += f"  {kat}: {fmt(total)}\n"

    if data:
        teks += "\n10 Transaksi Terakhir:\n"
        for r in data[-10:]:
            icon = "+" if r["Tipe"] == "MASUK" else "-"
            teks += f"{icon} {r['Deskripsi']}: {fmt(r['Nominal'])}\n"
    else:
        teks += "\nTidak ada transaksi."

    await update.message.reply_text(teks)


@owner_only
async def cmd_hari(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await laporan_tanggal(update, today_str(), "Hari Ini")


@owner_only
async def cmd_kemarin(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    kemarin = (now_tz() - dt.timedelta(days=1)).strftime("%Y-%m-%d")
    await laporan_tanggal(update, kemarin, "Kemarin")


# ──────────────────────────────────────────────
#  COMMAND: /bulan
# ──────────────────────────────────────────────
@owner_only
async def cmd_bulan(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sh      = get_sheet()
    ws      = get_ws(sh, "Transaksi")
    records = ws.get_all_records()
    bulan   = this_month()
    data    = [r for r in records if str(r["Tanggal"]).startswith(bulan)]

    masuk  = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "MASUK")
    keluar = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "KELUAR")

    kat_totals = {}
    for r in data:
        if r["Tipe"] == "KELUAR":
            kat = r.get("Kategori", "lain")
            kat_totals[kat] = kat_totals.get(kat, 0) + int(r["Nominal"])

    label = now_tz().strftime("%B %Y")
    teks  = f"Laporan Bulan {label}\n\n"
    teks += f"Total Pemasukan : {fmt(masuk)}\n"
    teks += f"Total Pengeluaran: {fmt(keluar)}\n"
    teks += f"Laba Bersih      : {fmt(masuk - keluar)}\n"
    teks += f"Total Transaksi  : {len(data)}\n"

    if kat_totals:
        teks += "\nRincian Pengeluaran:\n"
        for kat, total in sorted(kat_totals.items(), key=lambda x: -x[1]):
            persen = total / keluar * 100 if keluar > 0 else 0
            teks += f"  {kat}: {fmt(total)} ({persen:.1f}%)\n"

    await update.message.reply_text(teks)


# ──────────────────────────────────────────────
#  COMMAND: /labarugi
# ──────────────────────────────────────────────
@owner_only
async def cmd_labarugi(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sh      = get_sheet()
    ws      = get_ws(sh, "Transaksi")
    records = ws.get_all_records()
    bulan   = this_month()
    data    = [r for r in records if str(r["Tanggal"]).startswith(bulan)]

    penjualan    = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "MASUK" and r.get("Kategori") == "penjualan")
    masuk_lain   = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "MASUK" and r.get("Kategori") != "penjualan")
    total_masuk  = penjualan + masuk_lain
    bahan_baku   = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "KELUAR" and r.get("Kategori") == "bahan_baku")
    operasional  = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "KELUAR" and r.get("Kategori") == "operasional")
    gaji         = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "KELUAR" and r.get("Kategori") == "gaji")
    lain_keluar  = sum(int(r["Nominal"]) for r in data if r["Tipe"] == "KELUAR" and r.get("Kategori") not in ["bahan_baku","operasional","gaji"])
    total_keluar = bahan_baku + operasional + gaji + lain_keluar
    laba_bersih  = total_masuk - total_keluar

    label = now_tz().strftime("%B %Y")
    teks  = f"LAPORAN LABA RUGI - {label}\n"
    teks += "━━━━━━━━━━━━━━━━━━━━━\n\n"
    teks += "PENDAPATAN\n"
    teks += f"  Penjualan       : {fmt(penjualan)}\n"
    teks += f"  Pendapatan Lain : {fmt(masuk_lain)}\n"
    teks += f"  Total           : {fmt(total_masuk)}\n\n"
    teks += "BEBAN / BIAYA\n"
    teks += f"  Bahan Baku      : {fmt(bahan_baku)}\n"
    teks += f"  Operasional     : {fmt(operasional)}\n"
    teks += f"  Gaji            : {fmt(gaji)}\n"
    teks += f"  Lain-lain       : {fmt(lain_keluar)}\n"
    teks += f"  Total           : {fmt(total_keluar)}\n\n"
    teks += "━━━━━━━━━━━━━━━━━━━━━\n"
    teks += f"LABA BERSIH : {fmt(laba_bersih)}"

    await update.message.reply_text(teks)


# ──────────────────────────────────────────────
#  COMMAND: /prodhari & /prodbul
# ──────────────────────────────────────────────
@owner_only
async def cmd_prodhari(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sh      = get_sheet()
    ws      = get_ws(sh, "Produksi")
    records = ws.get_all_records()
    tgl     = today_str()
    data    = [r for r in records if r["Tanggal"] == tgl]

    if not data:
        await update.message.reply_text(f"Belum ada data produksi hari ini ({tgl}).")
        return

    prod_map = {}
    for r in data:
        prod_map[r["Produk"]] = prod_map.get(r["Produk"], 0) + int(r["Jumlah"])

    teks = f"Produksi Hari Ini ({tgl})\n\n"
    for prod, jml in sorted(prod_map.items(), key=lambda x: -x[1]):
        teks += f"  {prod}: {jml} pcs\n"
    teks += f"\nTotal: {sum(prod_map.values())} item"
    await update.message.reply_text(teks)


@owner_only
async def cmd_prodbul(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sh      = get_sheet()
    ws      = get_ws(sh, "Produksi")
    records = ws.get_all_records()
    bulan   = this_month()
    data    = [r for r in records if str(r["Tanggal"]).startswith(bulan)]

    if not data:
        await update.message.reply_text(f"Belum ada data produksi bulan ini ({bulan}).")
        return

    prod_map = {}
    for r in data:
        prod_map[r["Produk"]] = prod_map.get(r["Produk"], 0) + int(r["Jumlah"])

    label = now_tz().strftime("%B %Y")
    teks  = f"Rekap Produksi {label}\n\n"
    for prod, jml in sorted(prod_map.items(), key=lambda x: -x[1]):
        teks += f"  {prod}: {jml} pcs\n"
    teks += f"\nTotal: {sum(prod_map.values())} item\n"
    teks += f"Hari aktif: {len(set(r['Tanggal'] for r in data))} hari"
    await update.message.reply_text(teks)


# ──────────────────────────────────────────────
#  COMMAND: /hapus & /hapusprod
# ──────────────────────────────────────────────
@owner_only
async def cmd_hapus(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sh      = get_sheet()
    ws      = get_ws(sh, "Transaksi")
    records = ws.get_all_records()
    if not records:
        await update.message.reply_text("Tidak ada transaksi untuk dihapus.")
        return
    last = records[-1]
    ws.delete_rows(len(records) + 1)
    await update.message.reply_text(
        f"Transaksi terakhir dihapus:\n"
        f"{last['Deskripsi']}: {fmt(last['Nominal'])} ({last['Tipe']})"
    )


@owner_only
async def cmd_hapusprod(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sh      = get_sheet()
    ws      = get_ws(sh, "Produksi")
    records = ws.get_all_records()
    if not records:
        await update.message.reply_text("Tidak ada data produksi untuk dihapus.")
        return
    last = records[-1]
    ws.delete_rows(len(records) + 1)
    await update.message.reply_text(
        f"Produksi terakhir dihapus:\n"
        f"{last['Produk']}: {last['Jumlah']} {last['Satuan']}"
    )


# ──────────────────────────────────────────────
#  BUILD EXCEL
# ──────────────────────────────────────────────
def build_excel(tx_records, prod_records, label):
    wb   = openpyxl.Workbook()
    thin = Side(style="thin", color="DDDDDD")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # Sheet 1: Ringkasan
    ws1 = wb.active
    ws1.title = "Ringkasan"
    masuk  = sum(int(r["Nominal"]) for r in tx_records if r["Tipe"] == "MASUK")
    keluar = sum(int(r["Nominal"]) for r in tx_records if r["Tipe"] == "KELUAR")
    laba   = masuk - keluar

    ws1["A1"] = f"LAPORAN KEUANGAN - {label}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Dicetak: {now_tz().strftime('%d %B %Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, size=10, color="888888")
    ws1.append([])
    ws1.append(["Keterangan", "Nominal"])
    ws1.append(["Total Pemasukan", masuk])
    ws1.append(["Total Pengeluaran", keluar])
    ws1.append(["Laba Bersih", laba])
    ws1.append([])
    ws1.append(["PENGELUARAN PER KATEGORI", ""])

    kat_totals = {}
    for r in tx_records:
        if r["Tipe"] == "KELUAR":
            kat = r.get("Kategori", "lain")
            kat_totals[kat] = kat_totals.get(kat, 0) + int(r["Nominal"])
    for kat, tot in sorted(kat_totals.items(), key=lambda x: -x[1]):
        ws1.append([kat, tot])

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20

    # Sheet 2: Transaksi
    ws2   = wb.create_sheet("Transaksi")
    hdrs2 = ["No","Tanggal","Waktu","Tipe","Kategori","Deskripsi","Nominal","Sumber"]
    ws2.append(hdrs2)
    for col, h in enumerate(hdrs2, 1):
        c = ws2.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill("1E7A4A")
        c.alignment = Alignment(horizontal="center")

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill   = PatternFill("solid", fgColor="FFC7CE")

    for i, r in enumerate(tx_records, 1):
        ws2.append([i, r["Tanggal"], r["Waktu"], r["Tipe"], r.get("Kategori",""), r["Deskripsi"], int(r["Nominal"]), r.get("Sumber","")])
        rn = ws2.max_row
        ws2.cell(rn, 4).fill = green_fill if r["Tipe"] == "MASUK" else red_fill
        ws2.cell(rn, 7).number_format = '#,##0'
        ws2.cell(rn, 7).alignment = Alignment(horizontal="right")
        for col in range(1, len(hdrs2)+1):
            ws2.cell(rn, col).border = bdr
    auto_width(ws2)

    # Sheet 3: Produksi
    ws3   = wb.create_sheet("Produksi")
    hdrs3 = ["No","Tanggal","Waktu","Produk","Jumlah","Satuan","Catatan"]
    ws3.append(hdrs3)
    for col, h in enumerate(hdrs3, 1):
        c = ws3.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill("B7770D")
        c.alignment = Alignment(horizontal="center")

    prod_map = {}
    for i, r in enumerate(prod_records, 1):
        ws3.append([i, r["Tanggal"], r["Waktu"], r["Produk"], int(r["Jumlah"]), r.get("Satuan","pcs"), r.get("Catatan","")])
        rn = ws3.max_row
        ws3.cell(rn, 5).alignment = Alignment(horizontal="center")
        for col in range(1, len(hdrs3)+1):
            ws3.cell(rn, col).border = bdr
        prod_map[r["Produk"]] = prod_map.get(r["Produk"], 0) + int(r["Jumlah"])
    auto_width(ws3)

    # Sheet 4: Rekap Produksi
    ws4   = wb.create_sheet("Rekap Produksi")
    hdrs4 = ["Produk", "Total", "Satuan"]
    ws4.append(hdrs4)
    for col, h in enumerate(hdrs4, 1):
        c = ws4.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill("B7770D")

    for prod, jml in sorted(prod_map.items(), key=lambda x: -x[1]):
        ws4.append([prod, jml, "pcs"])
        ws4.cell(ws4.max_row, 2).alignment = Alignment(horizontal="center")

    if prod_map:
        chart = BarChart()
        chart.type   = "col"
        chart.title  = f"Produksi - {label}"
        chart.style  = 10
        chart.width  = 20
        chart.height = 14
        data_ref = Reference(ws4, min_col=2, min_row=1, max_row=ws4.max_row)
        cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws4.add_chart(chart, "E2")
    auto_width(ws4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────
#  COMMAND: /excel & /excelbulan
# ──────────────────────────────────────────────
@owner_only
async def cmd_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Membuat file Excel bulan ini...")
    await kirim_excel(update, this_month())


@owner_only
async def cmd_excelbulan(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    args = ctx.args
    if not args:
        await update.message.reply_text("Format: /excelbulan YYYY-MM\nContoh: /excelbulan 2025-01")
        return
    bulan = args[0].strip()
    if not re.match(r"^\d{4}-\d{2}$", bulan):
        await update.message.reply_text("Format salah. Gunakan: YYYY-MM")
        return
    await update.message.reply_text(f"Membuat file Excel untuk {bulan}...")
    await kirim_excel(update, bulan)


async def kirim_excel(update, bulan):
    try:
        sh       = get_sheet()
        tx_all   = get_ws(sh, "Transaksi").get_all_records()
        prod_all = get_ws(sh, "Produksi").get_all_records()
        tx_bul   = [r for r in tx_all   if str(r["Tanggal"]).startswith(bulan)]
        prod_bul = [r for r in prod_all if str(r["Tanggal"]).startswith(bulan)]

        if not tx_bul and not prod_bul:
            await update.message.reply_text(f"Tidak ada data untuk {bulan}.")
            return

        try:
            label = dt.datetime.strptime(bulan, "%Y-%m").strftime("%B %Y")
        except Exception:
            label = bulan

        buf  = build_excel(tx_bul, prod_bul, label)
        nama = f"Laporan_{bulan.replace('-','_')}.xlsx"
        await update.message.reply_document(
            document=buf,
            filename=nama,
            caption=f"Laporan {label}\n{len(tx_bul)} transaksi | {len(prod_bul)} entri produksi"
        )
    except Exception as e:
        await update.message.reply_text(f"Gagal buat Excel: {e}")
        logger.error(f"Excel error: {e}")


# ──────────────────────────────────────────────
#  LAPORAN HARIAN OTOMATIS
# ──────────────────────────────────────────────
async def kirim_laporan_harian(ctx: ContextTypes.DEFAULT_TYPE):
    try:
        sh        = get_sheet()
        tgl       = today_str()
        tx_all    = get_ws(sh, "Transaksi").get_all_records()
        prod_all  = get_ws(sh, "Produksi").get_all_records()
        tx_hari   = [r for r in tx_all   if r["Tanggal"] == tgl]
        prod_hari = [r for r in prod_all if r["Tanggal"] == tgl]

        masuk  = sum(int(r["Nominal"]) for r in tx_hari if r["Tipe"] == "MASUK")
        keluar = sum(int(r["Nominal"]) for r in tx_hari if r["Tipe"] == "KELUAR")

        prod_map = {}
        for r in prod_hari:
            prod_map[r["Produk"]] = prod_map.get(r["Produk"], 0) + int(r["Jumlah"])

        teks  = "LAPORAN HARIAN OTOMATIS\n"
        teks += f"Tanggal: {tgl}\n"
        teks += "━━━━━━━━━━━━━━━━━━━━━\n\n"
        teks += "KEUANGAN\n"
        teks += f"  Pemasukan  : {fmt(masuk)}\n"
        teks += f"  Pengeluaran: {fmt(keluar)}\n"
        teks += f"  Laba       : {fmt(masuk - keluar)}\n\n"

        if prod_map:
            teks += "PRODUKSI\n"
            for prod, jml in sorted(prod_map.items(), key=lambda x: -x[1]):
                teks += f"  {prod}: {jml} pcs\n"
            teks += f"  Total: {sum(prod_map.values())} item\n"
        else:
            teks += "PRODUKSI: Tidak ada data\n"

        teks += "\n━━━━━━━━━━━━━━━━━━━━━\n"
        teks += "Laporan otomatis"

        await ctx.bot.send_message(chat_id=OWNER_ID, text=teks)
    except Exception as e:
        logger.error(f"Laporan harian error: {e}")


# ──────────────────────────────────────────────
#  HANDLER PESAN UTAMA
# ──────────────────────────────────────────────
@owner_only
async def handle_message(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text  = update.message.text.strip()
    lower = text.lower()

    # PRODUKSI
    if lower.startswith("produksi") or lower.startswith("prod "):
        items = parse_produksi(text)
        if not items:
            await update.message.reply_text(
                "Format produksi tidak dikenali.\n\n"
                "Contoh:\n"
                "produksi roti coklat 50\n"
                "prod croissant 30 loyang, bolu 20"
            )
            return

        sh  = get_sheet()
        ws  = get_ws(sh, "Produksi")
        tgl = now_tz().strftime("%Y-%m-%d")
        wkt = now_tz().strftime("%H:%M")

        balasan = "Produksi Tercatat!\n\n"
        for item in items:
            tx_id = get_next_id(ws)
            ws.append_row([tx_id, tgl, wkt, item["produk"].title(), item["jumlah"], item["satuan"], ""])
            balasan += f"  {item['produk'].title()}: {item['jumlah']} {item['satuan']}\n"

        balasan += f"\nWaktu: {wkt} | {tgl}"
        await update.message.reply_text(balasan)
        return

    # TRANSAKSI
    hasil = parse_transaksi(text)
    if not hasil:
        await update.message.reply_text(
            "Format tidak dikenali.\n\n"
            "Contoh pengeluaran: beli tepung 50000\n"
            "Contoh pemasukan: + penjualan roti 200000\n\n"
            "Ketik /menu untuk bantuan lengkap."
        )
        return

    sh  = get_sheet()
    ensure_sheets(sh)
    ws  = get_ws(sh, "Transaksi")
    tgl = now_tz().strftime("%Y-%m-%d")
    wkt = now_tz().strftime("%H:%M")
    kat = auto_kategori(hasil["deskripsi"], hasil["tipe"])
    tid = get_next_id(ws)

    ws.append_row([
        tid, tgl, wkt,
        hasil["tipe"], kat,
        hasil["deskripsi"].capitalize(),
        hasil["nominal"],
        "telegram"
    ])

    # Hitung total hari ini
    semua    = ws.get_all_records()
    hari_ini = [r for r in semua if r["Tanggal"] == tgl]
    t_masuk  = sum(int(r["Nominal"]) for r in hari_ini if r["Tipe"] == "MASUK")
    t_keluar = sum(int(r["Nominal"]) for r in hari_ini if r["Tipe"] == "KELUAR")
    laba     = t_masuk - t_keluar
    jml_tx   = len(hari_ini)

    tipe_label = "Pemasukan" if hasil["tipe"] == "MASUK" else "Pengeluaran"
    icon       = "[MASUK]" if hasil["tipe"] == "MASUK" else "[KELUAR]"

    await update.message.reply_text(
        f"{icon} {tipe_label} Tercatat!\n"
        f"━━━━━━━━━━━━━━━━━\n"
        f"Deskripsi : {hasil['deskripsi'].capitalize()}\n"
        f"Nominal   : {fmt(hasil['nominal'])}\n"
        f"Kategori  : {kat}\n"
        f"Waktu     : {wkt} | {tgl}\n"
        f"━━━━━━━━━━━━━━━━━\n"
        f"Rekap Hari Ini ({jml_tx} transaksi)\n"
        f"Total Masuk  : {fmt(t_masuk)}\n"
        f"Total Keluar : {fmt(t_keluar)}\n"
        f"Laba Hari Ini: {fmt(laba)}"
    )


# ──────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────
def main():
    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start",      cmd_start))
    app.add_handler(CommandHandler("menu",       cmd_menu))
    app.add_handler(CommandHandler("saldo",      cmd_saldo))
    app.add_handler(CommandHandler("hari",       cmd_hari))
    app.add_handler(CommandHandler("kemarin",    cmd_kemarin))
    app.add_handler(CommandHandler("bulan",      cmd_bulan))
    app.add_handler(CommandHandler("labarugi",   cmd_labarugi))
    app.add_handler(CommandHandler("prodhari",   cmd_prodhari))
    app.add_handler(CommandHandler("prodbul",    cmd_prodbul))
    app.add_handler(CommandHandler("excel",      cmd_excel))
    app.add_handler(CommandHandler("excelbulan", cmd_excelbulan))
    app.add_handler(CommandHandler("hapus",      cmd_hapus))
    app.add_handler(CommandHandler("hapusprod",  cmd_hapusprod))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    waktu_laporan = dt.time(hour=JAM_LAPORAN, minute=0, second=0, tzinfo=TZ)
    app.job_queue.run_daily(kirim_laporan_harian, time=waktu_laporan, name="laporan_harian")

    logger.info("Bot Keuangan & Produksi berjalan...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
